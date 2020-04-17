from tkinter import *
import tkinter.filedialog
import requests
from bs4 import BeautifulSoup
import xlrd
import xlwt
#判断一段文本中是否包含简体中文
import re
# 写excel
from openpyxl import Workbook

pattern = re.compile('[0-9]+')
zhmodel = re.compile(u'[\u4e00-\u9fa5]')  #检查中文

TasteName = []
TasteEnergy = []


def getHTML(url):
    headers = {
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)"
                      " AppleWebKit/537.36 (KHTML, like Gecko) Chrome/77.0.3865.120 Safari/537.36"
    }
    try:
        html = requests.get(url, headers=headers, timeout=20)
        html.raise_for_status()
        html.encoding = html.apparent_encoding
        return html.text
    except ConnectionError and TimeoutError:
        getHTML(url)
    except:
        return "Error"


def paserHtml(html):
    soup = BeautifulSoup(html, "lxml")
    texts = soup.find_all("font", style="font-size: 14px")
    for text in texts:
        Foodname = text.find_all(text=True)
        for food in Foodname:
            strTemp = food.strip()
            if strTemp != "" and strTemp != "减肥":
                if zhmodel.search(strTemp):
                    if strTemp != "食品名称" and strTemp != "热量(大卡)/可食部分(克)" and strTemp != "热量(大卡/可食部(克)":
                        TasteName.append(strTemp)
                else:
                    TasteEnergy.append(strTemp)

    '''
    tbodys = soup.find_all("tbody")
    for tbody in tbodys:
        trs = tbody.find_all("tr")
        for tr in trs:
            tds = tr.find_all("td")
            for td in tds:
                ps = td.find_all("p")
                for p in ps:
                    texts = td.find_all("font", style="font-size: 14px")
                    for text in texts:
                        Foodname = text.find_all(text=True)
                        for food in Foodname:
                            strTemp = food.strip()
                            if zhmodel.search(strTemp):
                                if strTemp != "食品名称" and strTemp != "热量(大卡)/可食部分(克)":
                                    TasteName.append(strTemp)
                            else:
                                TasteEnergy.append(strTemp)
    '''


def write_excel():
    MyExcel = Workbook()
    Mysheet = MyExcel.active
    Mysheet.title = '食物'

    Mysheet.cell(row=1, column=1, value="食品名称")
    Mysheet.cell(row=1, column=2, value="热量(大卡)/可食部分(克)")

    row = 2
    for x in TasteName:
        Mysheet.cell(row=row, column=1, value=x)
        row += 1
    row = 2
    for y in TasteEnergy:
        Mysheet.cell(row=row, column=2, value=y)
        row += 1
    MyExcel.save('食物.xlsx')


if __name__ == '__main__':
    url = [
        "http://www.shoumm.com/",
    ]
    for x in url:
        html = getHTML(x)
        paserHtml(html)

    write_excel()
    print("成功！")








