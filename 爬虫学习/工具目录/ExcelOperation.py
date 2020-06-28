from openpyxl import Workbook
from openpyxl import load_workbook


class ExcelOperation(object):
    def __init__(self, name):
        self.excel_name = name
        self.excel_workbook = self.load_excel()

    def __del__(self):
        try:
            self.excel_workbook.save(self.excel_name)
        except Exception as e:
            print(e)

    def load_excel(self):
        try:
            wb = load_workbook(self.excel_name)
            return wb
        except FileNotFoundError as e:
            wb = Workbook()  # 创建一个工作簿
            return wb

    def create_sheet(self, sheet_name):
        self.excel_workbook.create_sheet(sheet_name)  # 创建一个工作表

    def write_in_cell(self, row, col, value):
        ws = self.excel_workbook.active
        ws.cell(row, col, value)
